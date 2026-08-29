from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _format_value(value) -> str:
    if value is None or value == "":
        return "-"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}"
    return str(value)


def _pdf_header(canvas, doc, title: str) -> None:
    canvas.saveState()
    canvas.setAuthor("Andres Felipe Figueroa")
    canvas.setCreator("Portal de atención al cliente")
    canvas.setTitle(title)
    canvas.setSubject("Reporte mensual de solicitudes")
    canvas.setFillColor(colors.HexColor("#0f172a"))
    canvas.rect(0, A4[1] - 18 * mm, A4[0], 18 * mm, fill=1, stroke=0)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 12)
    canvas.drawString(18 * mm, A4[1] - 11 * mm, title)
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(A4[0] - 18 * mm, A4[1] - 11 * mm, "Portal web de atención al cliente")
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.setFont("Helvetica", 8)
    canvas.drawString(18 * mm, 10 * mm, f"Página {canvas.getPageNumber()}")
    canvas.restoreState()


def generate_pdf_report(report: dict, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=26 * mm,
        bottomMargin=16 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="ReportTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=10,
        )
    )
    styles.add(
        ParagraphStyle(
            name="ReportSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#475569"),
            spaceAfter=12,
        )
    )
    styles.add(
        ParagraphStyle(
            name="ReportSection",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=14,
            textColor=colors.HexColor("#0f172a"),
            spaceBefore=8,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="ReportBody",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=11,
            textColor=colors.HexColor("#334155"),
        )
    )

    story = []
    story.append(Paragraph(f"Reporte mensual {report['label']}", styles["ReportTitle"]))
    story.append(
        Paragraph(
            "Resumen ejecutivo del comportamiento de las solicitudes registradas en el portal.",
            styles["ReportSubtitle"],
        )
    )

    summary_rows = [
        ["Métrica", "Valor", "Métrica", "Valor"],
        ["Total solicitudes", _format_value(report["summary"].get("total_tickets")), "Cerradas", _format_value(report["summary"].get("closed_tickets"))],
        ["Abiertas", _format_value(report["summary"].get("open_tickets")), "Pendientes del cliente", _format_value(report["summary"].get("pending_client"))],
        ["Promedio atención (h)", _format_value(report["summary"].get("avg_resolution_hours")), "Satisfacción", _format_value(report["summary"].get("satisfaction_avg"))],
    ]
    summary_table = Table(summary_rows, colWidths=[45 * mm, 35 * mm, 45 * mm, 35 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f8fafc"), colors.HexColor("#e2e8f0")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 8))

    def add_section(title: str, rows: list[list[str]], widths: list[int]) -> None:
        story.append(Paragraph(title, styles["ReportSection"]))
        table = Table(rows, colWidths=widths)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0ea5e9")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 8))

    status_rows = [["Estado", "Total"]]
    for row in report.get("by_status", []):
        status_rows.append([row["status"], _format_value(row["total"])])
    if len(status_rows) == 1:
        status_rows.append(["Sin datos", "0"])
    add_section("Distribución por estado", status_rows, [120 * mm, 40 * mm])

    category_rows = [["Categoría", "Total"]]
    for row in report.get("by_category", []):
        category_rows.append([row["name"], _format_value(row["total"])])
    if len(category_rows) == 1:
        category_rows.append(["Sin datos", "0"])
    add_section("Distribución por categoría", category_rows, [120 * mm, 40 * mm])

    employee_rows = [["Empleado", "Total"]]
    for row in report.get("by_employee", []):
        employee_rows.append([row["name"], _format_value(row["total"])])
    if len(employee_rows) == 1:
        employee_rows.append(["Sin datos", "0"])
    add_section("Carga por empleado", employee_rows, [120 * mm, 40 * mm])

    ticket_rows = [["Ticket", "Cliente", "Estado", "Prioridad", "Categoría", "Actualizado"]]
    for ticket in report.get("tickets", [])[:15]:
        ticket_rows.append(
            [
                ticket["code"],
                ticket["client_name"],
                ticket["status"],
                ticket["priority"],
                ticket["category_name"],
                ticket["updated_at"][:16] if ticket.get("updated_at") else "-",
            ]
        )
    if len(ticket_rows) == 1:
        ticket_rows.append(["Sin solicitudes", "-", "-", "-", "-", "-"])
    add_section(
        "Solicitudes del período",
        ticket_rows,
        [30 * mm, 40 * mm, 25 * mm, 22 * mm, 35 * mm, 25 * mm],
    )

    doc.build(story, onFirstPage=lambda canvas, doc: _pdf_header(canvas, doc, f"Reporte {report['label']}"), onLaterPages=lambda canvas, doc: _pdf_header(canvas, doc, f"Reporte {report['label']}"))
    return output_path


def generate_excel_report(report: dict, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    title_fill = PatternFill("solid", fgColor="0F172A")
    accent_fill = PatternFill("solid", fgColor="0EA5E9")
    soft_fill = PatternFill("solid", fgColor="E2E8F0")
    header_font = Font(color="FFFFFF", bold=True)
    strong_font = Font(bold=True, color="0F172A")

    ws["A1"] = f"Reporte mensual {report['label']}"
    ws["A1"].font = Font(size=16, bold=True, color="0F172A")
    ws["A2"] = "Portal web de atención al cliente"
    ws["A2"].font = Font(size=10, color="475569")

    summary_rows = [
        ("Total solicitudes", report["summary"].get("total_tickets")),
        ("Cerradas", report["summary"].get("closed_tickets")),
        ("Abiertas", report["summary"].get("open_tickets")),
        ("Pendientes del cliente", report["summary"].get("pending_client")),
        ("Promedio atención (h)", report["summary"].get("avg_resolution_hours")),
        ("Satisfacción", report["summary"].get("satisfaction_avg")),
    ]

    row = 4
    for index, (label, value) in enumerate(summary_rows, start=0):
        col = 1 if index % 2 == 0 else 4
        if index % 2 == 0 and index > 0:
            row += 1
        ws.cell(row=row, column=col, value=label).font = strong_font
        ws.cell(row=row, column=col + 1, value=value)
        ws.cell(row=row, column=col).fill = soft_fill
        ws.cell(row=row, column=col + 1).fill = soft_fill

    def write_table(sheet_name: str, headers: list[str], rows: list[list[str]]) -> None:
        sheet = wb.create_sheet(sheet_name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = title_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_values in rows or [["Sin datos" if headers else "Sin datos"]]:
            sheet.append(row_values)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell_text = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_text))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
        return sheet

    status_rows = [[row["status"], row["total"]] for row in report.get("by_status", [])] or [["Sin datos", 0]]
    category_rows = [[row["name"], row["total"]] for row in report.get("by_category", [])] or [["Sin datos", 0]]
    employee_rows = [[row["name"], row["total"]] for row in report.get("by_employee", [])] or [["Sin datos", 0]]
    ticket_rows = [
        [t["code"], t["client_name"], t["status"], t["priority"], t["category_name"], t["updated_at"]]
        for t in report.get("tickets", [])
    ] or [["Sin datos", "-", "-", "-", "-", "-"]]

    write_table("Estados", ["Estado", "Total"], status_rows)
    write_table("Categorias", ["Categoría", "Total"], category_rows)
    write_table("Empleados", ["Empleado", "Total"], employee_rows)
    write_table("Solicitudes", ["Ticket", "Cliente", "Estado", "Prioridad", "Categoría", "Actualizado"], ticket_rows)

    summary_sheet = wb["Resumen"]
    summary_sheet.freeze_panes = "A4"
    for col in ("A", "B", "C", "D", "E", "F"):
        summary_sheet.column_dimensions[col].width = 22

    wb.save(output_path)
    return output_path
